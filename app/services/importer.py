from __future__ import annotations

import io
import logging
import re
import zipfile
from dataclasses import dataclass

import httpx
from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.db.models import SimpleTabulationCode


logger = logging.getLogger(__name__)

REQUIRED_HEADERS = {
    "Code",
    "Title",
    "ClassKind",
    "DepthInKind",
    "ChapterNo",
}
GROUPING_COLUMNS = ["Grouping5", "Grouping4", "Grouping3", "Grouping2", "Grouping1"]
TITLE_PREFIX_PATTERN = re.compile(r"^\s*(?:-\s*)+")


class ImporterError(RuntimeError):
    pass


@dataclass(slots=True)
class ImportSummary:
    source_name: str
    spreadsheet_name: str
    sheet_name: str
    imported_rows: int
    inserted_rows: int
    updated_rows: int
    replace_mode: bool

    def to_dict(self) -> dict[str, str | int | bool]:
        return {
            "source_name": self.source_name,
            "spreadsheet_name": self.spreadsheet_name,
            "sheet_name": self.sheet_name,
            "imported_rows": self.imported_rows,
            "inserted_rows": self.inserted_rows,
            "updated_rows": self.updated_rows,
            "replace_mode": self.replace_mode,
        }


class SimpleTabulationImporter:
    def __init__(self, session: Session) -> None:
        self.session = session

    def import_from_url(self, zip_url: str, replace: bool = True) -> ImportSummary:
        zip_bytes = self._download_zip(zip_url)
        return self.import_from_zip_bytes(
            zip_bytes=zip_bytes,
            source_name=zip_url,
            replace=replace,
        )

    def import_from_zip_bytes(
        self,
        zip_bytes: bytes,
        source_name: str,
        replace: bool = True,
    ) -> ImportSummary:
        try:
            archive = zipfile.ZipFile(io.BytesIO(zip_bytes))
        except zipfile.BadZipFile as exc:
            logger.exception("Import error: invalid ZIP archive")
            raise ImporterError("The downloaded file is not a valid ZIP archive.") from exc

        workbook_name = self._find_workbook_name(archive)
        sheet_name, records = self._extract_records(archive, workbook_name)
        return self._persist_records(
            source_name=source_name,
            spreadsheet_name=workbook_name,
            sheet_name=sheet_name,
            records=records,
            replace=replace,
        )

    @staticmethod
    def _download_zip(zip_url: str) -> bytes:
        try:
            with httpx.Client(timeout=90.0, follow_redirects=True) as client:
                response = client.get(zip_url)
                response.raise_for_status()
                return response.content
        except httpx.HTTPError as exc:
            logger.exception("Import error while downloading %s", zip_url)
            raise ImporterError(f"Unable to download WHO ZIP file from {zip_url}") from exc

    @staticmethod
    def _find_workbook_name(archive: zipfile.ZipFile) -> str:
        workbook_names = [
            name
            for name in archive.namelist()
            if name.lower().endswith((".xlsx", ".xlsm"))
        ]
        if not workbook_names:
            raise ImporterError("ZIP structure changed: no spreadsheet file was found.")
        if len(workbook_names) == 1:
            return workbook_names[0]

        for name in workbook_names:
            if "simpletabulation" in name.lower():
                return name
        raise ImporterError(
            "ZIP structure changed: multiple spreadsheets were found and none matched the expected name."
        )

    def _extract_records(
        self,
        archive: zipfile.ZipFile,
        workbook_name: str,
    ) -> tuple[str, list[dict[str, object]]]:
        workbook_bytes = archive.read(workbook_name)

        try:
            workbook = load_workbook(
                filename=io.BytesIO(workbook_bytes),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            logger.exception("Import error while opening workbook %s", workbook_name)
            raise ImporterError("Unable to open the WHO spreadsheet.") from exc

        try:
            sheet = self._find_target_sheet(workbook)
            headers = self._extract_headers(sheet)
            records = self._build_records(sheet, headers)
            return sheet.title, records
        finally:
            workbook.close()

    def _find_target_sheet(self, workbook):
        for sheet in workbook.worksheets:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row:
                continue
            headers = {self._normalize_header(value) for value in first_row if value is not None}
            if REQUIRED_HEADERS.issubset(headers):
                return sheet

        raise ImporterError(
            "Spreadsheet structure changed: no worksheet with the required headers was found."
        )

    def _extract_headers(self, sheet) -> list[str]:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ImporterError("Spreadsheet is missing its header row.")

        headers = [self._normalize_header(value) for value in header_row]
        if not REQUIRED_HEADERS.issubset({header for header in headers if header}):
            raise ImporterError(
                "Spreadsheet structure changed: required columns are missing from the worksheet."
            )
        return headers

    def _build_records(self, sheet, headers: list[str]) -> list[dict[str, object]]:
        records: list[dict[str, object]] = []
        last_code_by_depth: dict[int, str] = {}

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_row = {
                header: self._normalize_cell(value)
                for header, value in zip(headers, row, strict=False)
                if header
            }

            code = self._clean_code(raw_row.get("Code"))
            if not code:
                continue

            depth = self._to_int(raw_row.get("DepthInKind"))
            parent_code = last_code_by_depth.get(depth - 1) if depth > 1 else None
            last_code_by_depth = {
                existing_depth: existing_code
                for existing_depth, existing_code in last_code_by_depth.items()
                if existing_depth < depth
            }
            last_code_by_depth[depth] = code

            title = self._clean_title(raw_row.get("Title"))
            records.append(
                {
                    "code": code,
                    "title": title,
                    "parent_code": parent_code,
                    "is_extension": self._is_extension(code, raw_row.get("ChapterNo")),
                    "sort_key": row_number,
                    "chapter_or_group": self._resolve_chapter_or_group(raw_row),
                    "raw_row_json": raw_row,
                }
            )

        if not records:
            raise ImporterError("Spreadsheet structure changed: no ICD-11 codes were extracted.")

        return records

    def _persist_records(
        self,
        source_name: str,
        spreadsheet_name: str,
        sheet_name: str,
        records: list[dict[str, object]],
        replace: bool,
    ) -> ImportSummary:
        inserted_rows = 0
        updated_rows = 0

        try:
            if replace:
                self.session.execute(delete(SimpleTabulationCode))

            existing_by_code = {}
            if not replace:
                existing_rows = self.session.scalars(select(SimpleTabulationCode)).all()
                existing_by_code = {row.code: row for row in existing_rows}

            for record in records:
                code = str(record["code"])
                existing = existing_by_code.get(code)
                if existing:
                    existing.title = str(record["title"])
                    existing.parent_code = record["parent_code"]
                    existing.is_extension = bool(record["is_extension"])
                    existing.sort_key = int(record["sort_key"])
                    existing.chapter_or_group = record["chapter_or_group"]
                    existing.raw_row_json = dict(record["raw_row_json"])
                    updated_rows += 1
                    continue

                self.session.add(SimpleTabulationCode(**record))
                inserted_rows += 1

            self.session.commit()
            logger.info(
                "Imported %s rows from %s (%s / %s)",
                len(records),
                source_name,
                spreadsheet_name,
                sheet_name,
            )
        except Exception as exc:
            self.session.rollback()
            logger.exception("Import error while persisting WHO Simple Tabulation data")
            raise ImporterError("Unable to persist imported WHO reference rows into SQLite.") from exc

        return ImportSummary(
            source_name=source_name,
            spreadsheet_name=spreadsheet_name,
            sheet_name=sheet_name,
            imported_rows=len(records),
            inserted_rows=inserted_rows,
            updated_rows=updated_rows,
            replace_mode=replace,
        )

    @staticmethod
    def _normalize_header(value: object) -> str:
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _normalize_cell(value: object) -> object:
        if isinstance(value, str):
            return value.strip()
        return value

    @staticmethod
    def _clean_code(value: object) -> str | None:
        if value is None:
            return None
        cleaned = str(value).strip().upper()
        return cleaned or None

    @staticmethod
    def _clean_title(value: object) -> str:
        if value is None:
            return ""
        return TITLE_PREFIX_PATTERN.sub("", str(value).strip()).strip()

    @staticmethod
    def _to_int(value: object) -> int:
        if value in (None, ""):
            return 0
        return int(value)

    @staticmethod
    def _is_extension(code: str, chapter_value: object) -> bool:
        chapter = str(chapter_value).strip().upper() if chapter_value is not None else ""
        return code.startswith("X") or chapter == "X"

    @staticmethod
    def _resolve_chapter_or_group(raw_row: dict[str, object]) -> str | None:
        for grouping in GROUPING_COLUMNS:
            value = raw_row.get(grouping)
            if value:
                return str(value)

        chapter = raw_row.get("ChapterNo")
        if chapter:
            return str(chapter)
        return None

